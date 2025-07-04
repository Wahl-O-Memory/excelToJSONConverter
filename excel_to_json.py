import json
import os
import sys
from openpyxl import load_workbook

def convert_excel_to_json(excel_path, output_dir):
    wb = load_workbook(excel_path)
    ws = wb.active  # Use first sheet

    # Extract metadata (rows 1-3)
    election_data = {
        "name": ws.cell(row=1, column=2).value,
        "electionUId": ws.cell(row=2, column=2).value,
        "description": ws.cell(row=3, column=2).value,
    }

    # Extract parties (rows 5-9)
    parties = []
    for col in range(2, ws.max_column + 1):
        parties.append({
            "partyShortName": ws.cell(row=5, column=col).value,
            "partyLongName": ws.cell(row=6, column=col).value,
            "partyIsGoverning": ws.cell(row=7, column=col).value == "true",
            "partyUID": ws.cell(row=8, column=col).value,
            "partyHasAnswered":ws.cell(row=9, column=col).value,
            "partyLogo": ws.cell(row=10, column=col).value,
        })
    election_data["parties"] = parties

    # Extract questions (rows 11+)
    questions = []
    current_row = 12
    while current_row <= ws.max_row:
        if ws.cell(row=current_row, column=1).value:  # Question row
            question_text = ws.cell(row=current_row, column=1).value
            answers = []
            for party_uid, col in enumerate(range(2, ws.max_column + 1)):
                answers.append({
                    "party": party_uid,
                    "support": ws.cell(row=current_row + 1, column=col).value,
                    "answer": ws.cell(row=current_row, column=col).value,
                })
            questions.append({
                "questionText": question_text,
                "answers": answers,
            })
            current_row += 2  # Skip answer rows
        else:
            current_row += 1
    election_data["questions"] = questions

    # Save JSON
    output_path = os.path.join(output_dir, election_data["electionUId"]+".json")
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(election_data, f, indent=2, ensure_ascii=False)
    print(f"JSON saved to: {output_path}")

if __name__ == "__main__":
    # Drag-and-drop support: Check if file was dropped onto .exe
    if len(sys.argv) > 1:
        input_excel = sys.argv[1]
        output_dir = os.path.dirname(input_excel)
        convert_excel_to_json(input_excel, output_dir)
    else:
        # If no file dragged, process all .xlsx files in the same folder as the .exe
        exe_dir = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(__file__)
        for file in os.listdir(exe_dir):
            if file.endswith(".xlsx"):
                convert_excel_to_json(os.path.join(exe_dir, file), exe_dir)
        input("Press Enter to exit...")